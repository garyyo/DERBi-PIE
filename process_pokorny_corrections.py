import glob
import json
import os
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

import gpt_functions
from find_missing_pokorny_gpt import get_web_lrc_lineup_corrected, get_web_from_lrc, augment_material, format_gpt_prompt_new2
from generate_pokorny_db_data import load_abbreviations_df, remove_html_tags_from_text

corrections_needed = {
    "Additional Pokorny Forms 0.xlsx": ["", ""],
    "Additional Pokorny Forms 1.xlsx": ["", ""],
    "Additional Pokorny Forms 2.xlsx": ["dheu-4, dheu̯ə- (vermutlich: dhu̯ē-, vgl. die Erw. dhu̯ē-k-, dhu̯ē̆-s-)", ""],
}


def mutate_font(font, **kwargs):
    # some common longhand
    translation_dict = {
        "bold": "b",
        "underline": "u",
        "italic": "i",
    }
    new_font = copy(font)
    for key, value in kwargs.items():
        new_font.__dict__[translation_dict.get(key, key)] = value
    return new_font


# todo: make sure I use forward slashes as I have assumed in the code here.
def un_italicize_cell(cell):
    # if the entire cell is italicized then we only put italics around the whole thing.
    if cell.font.italic:
        # weird roundabout way to set the italics to false but apparently its required
        cell.font = mutate_font(cell.font, italic=False)

        if isinstance(cell.value, CellRichText):
            for block_count, text_block in enumerate(cell.value):
                # make sure it is de-italicized
                if isinstance(text_block, TextBlock):
                    if text_block.font.italic:
                        text_block.font = mutate_font(text_block.font, italic=False)
                # add in my italics marker "/" to the start and end
                if block_count == 0:
                    if isinstance(text_block, TextBlock):
                        text_block.text = f"/{text_block.text}"
                    else:
                        cell.value[block_count] = f"/{text_block}"
                if block_count == (len(cell.value) - 1):
                    if isinstance(text_block, TextBlock):
                        text_block.text = f"{text_block.text}/"
                    else:
                        cell.value[block_count] = f"{text_block}/"
        elif isinstance(cell.value, str):
            cell.value = f"/{cell.value}/"
        elif cell.value is not None:
            breakpoint()
    # otherwise we do it a bunch of times.
    elif isinstance(cell.value, CellRichText):
        for text_block in cell.value:
            # ensure it's a text block, and not a plain string, and that it is italicized
            if isinstance(text_block, TextBlock) and text_block.font.italic:
                # mark it and remove the italics
                text_block.font = mutate_font(text_block.font, italic=False)
                text_block.text = f"/{text_block.text}/"


def fix_italics(filepath='data_pokorny/additional_pokorny_corrections/Additional Pokorny Forms 0.xlsx'):
    # Load the workbook
    workbook = load_workbook(filepath, rich_text=True)

    # Assume you're working with the first sheet, change as needed
    sheet = workbook.active

    # Iterate through all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            un_italicize_cell(cell)

    # Close the workbook when done
    path, ext = os.path.splitext(filepath)
    # workbook.save(filepath)
    new_filepath = f"{path}_italic_fix{ext}"
    workbook.save(new_filepath)
    workbook.close()
    return new_filepath


def extract_bad_entries(filepath):
    workbook = load_workbook(filepath, rich_text=True)
    sheet = workbook.active
    rerun_entry_ids = set()
    delete_entry_ids = set()
    normal_colors = ["00000000", "FFFFFF00", "FFFFFFFF"]
    error_color = "FFFF0000"
    redo_colors = ["FFFF9900"]

    for row in sheet.iter_rows():
        for cell in row:
            # if we find a color that has never been seen
            if cell.fill.start_color.rgb not in normal_colors + [error_color] + redo_colors:
                print(f"New cell color '{cell.fill.start_color.rgb}' in cell {cell}, with value '{cell.value}'")
                # breakpoint()

            # if we find a color marking deletion
            elif cell.fill.start_color.rgb == error_color:
                print("found error")
                if row[1].value is None or row[2].value is None:
                    print(f"Root entry highlighted, but no root found in cell {cell}, please correct and rerun")
                    # breakpoint()
                delete_entry_ids.add((row[1].value, row[2].value))
                break

            # if we find a color marking missing entries
            elif cell.fill.start_color.rgb in redo_colors:
                print("found redo")
                if row[1].value is None or row[2].value is None:
                    print(f"Root entry highlighted, but no root found in cell {cell}, please correct and rerun")
                    # breakpoint()
                rerun_entry_ids.add((row[1].value, row[2].value))
                break

    if (None, None) in rerun_entry_ids:
        rerun_entry_ids.remove((None, None))
    return rerun_entry_ids, delete_entry_ids


# the single most hacked together function I have ever written. If you are trying to read it, I am sorry.
def redo_pokorny(rerun_entry_ids, filepath):
    # we have the entries that we need to do, but we need all the other information about them
    dfs = {os.path.splitext(os.path.basename(df_file))[0]: pd.read_pickle(df_file) for df_file in glob.glob("data_pokorny/table_dumps/*.df")}
    abbreviation_data = load_abbreviations_df()
    line_up_df = get_web_lrc_lineup_corrected()

    redo_entries = dfs['lex_etyma'][dfs['lex_etyma'].entry.isin([id[1] for id in rerun_entry_ids])]
    output_dfs = []

    headers = ['abbr', 'reflex', 'meaning', 'notes']

    with open("data_pokorny/pokorny_scraped.json", "r", encoding="utf-8") as fp:
        scraped_pokorny = json.load(fp)

    # turn scraped pokorny into a dictionary
    scraped_pokorny_dict = {
        ", ".join(entry["root"]): entry
        for entry in scraped_pokorny
    }

    # if there are too many tokens (for which a character limit is an approximation for) we found that gpt3.5 just gets confused, even if it supports that many
    # so this char count limit is later used to make sure that any prompt over that limit goes to gpt4
    gpt4_char_threshold = 5000

    # just some data manipulation to make things look right
    entries_to_be_processed = list(enumerate(redo_entries.iterrows()))

    # colors to help print things out in a way that its easy to see
    header_color = "\033[92m"
    end_color = "\033[0m"

    for counter, (i, row) in entries_to_be_processed:
        # gather some data necessary for the prompt
        texas_root = remove_html_tags_from_text(row.entry).strip("\n")

        web_key = get_web_from_lrc(line_up_df, texas_root)
        web_entry = scraped_pokorny_dict.get(web_key, None)

        web_root = web_entry["root"]
        gloss = remove_html_tags_from_text(row.gloss)

        # gather material
        material = "\n".join(web_entry['Material']).replace("`", "'")

        # for some length threshold, if its under it, we try to run as normal
        if len(material) < 5000:
            materials = [material]
        # if over we try to run a bunch of smaller versions splitting the material based on the lines in the material field
        else:
            materials = []
            cumulative_line = ""
            cumulative_len = 0
            for line in web_entry['Material']:
                cumulative_len += len(line)
                cumulative_line += "\n" + line
                # we try to keep the line lengths under the gpt4 threshold, but this won't always work
                if cumulative_len > (gpt4_char_threshold - 3000):
                    materials.append(cumulative_line.strip())
                    cumulative_line = ""
                    cumulative_len = 0
            if cumulative_len != 0:
                materials.append(cumulative_line.strip())

        # here we actually start forming the prompt and running
        for mat_num, mat in enumerate(materials):
            # get progress info
            material_progress = f"{mat_num+1}/{len(materials)}"
            # get which abbreviations are *probably* used in this, it will err on the side of more abbreviations
            abbreviations_used = augment_material(mat, abbreviation_data)

            # form the actual prompt
            prompt = format_gpt_prompt_new2(texas_root, gloss, mat, abbreviations_used, headers)

            # we try to cut down to only use gpt3.5, but its difficult to plan that out with 100% accuracy,
            # so this switches to gpt4 in case we go over the limit for some reason
            model = "gpt-4" if len(prompt) > gpt4_char_threshold else "gpt-3.5-turbo"

            # check if it's valid before trying to grab the cached response
            if len(prompt) > 15000:
                print(f"{header_color}---=== Too long for {web_root} {material_progress} ===---{end_color}")
                # breakpoint()
            elif mat.strip() == "":
                print(f"{header_color}---=== No material for {web_root} {material_progress} ===---{end_color}")
                # breakpoint()
            else:
                attempt_failed = False
                for attempt in range(4):
                    try:
                        bypass_cache = (attempt != 0)
                        messages, _, digest = gpt_functions.get_cached([prompt], model=model)

                        if messages is None or bypass_cache:
                            print(f"{header_color}---=== Cached Missing for {web_root} {material_progress}! ===---{end_color}")
                            # for the sake of not wasting money, a gpt-4 request must be manually OK'd by having the user continue from the breakpoint.
                            # not ideal for users, but money.
                            if model == "gpt-4":
                                breakpoint()
                            messages, _ = gpt_functions.query_gpt(
                                [prompt],
                                model=model,
                                note=f"{counter}/{len(entries_to_be_processed)} {texas_root}{' '*15}"[:15],
                                bypass_cache=bypass_cache
                            )
                        else:
                            print(f"{header_color}---=== Found Cached for {web_root} {material_progress} ===---{end_color} {digest}")

                        content = gpt_functions.get_last_content(messages)
                        response = gpt_functions.extract_code_block(content)

                        # check if the first line is not valid csv and exclude it if it's not. It is likely to be marking what language the code block is.
                        if "," not in response.split("\n")[0]:
                            response = "\n".join(response.split("\n")[1:])

                        df = gpt_functions.csv_to_df(response, headers)

                        df["root"] = texas_root
                        df["web_root"] = web_root[-1]
                        output_dfs.append(df)
                        attempt_failed = False
                        break
                    except (pd.errors.ParserError, ) as err:
                        attempt_failed = True
                        print("\033[93mParse Error\033[0m")
                        continue
                if attempt_failed:
                    # breakpoint()
                    exit()
    if len(output_dfs) > 0:
        combined_df = pd.concat(output_dfs)
        combined_df = combined_df[['web_root', 'root'] + headers]
        os.path.basename(filepath)
        combined_df.to_csv(f"data_pokorny/additional_pokorny_corrections/rerun/{os.path.splitext(os.path.basename(filepath))[0]}.csv")
    pass


def main():
    filepath = 'data_pokorny/additional_pokorny_corrections/Additional Pokorny Forms 2.xlsx'
    # filepath_italic = fix_italics(filepath)
    # rerun_entry_ids, delete_entry_ids = extract_bad_entries(filepath_italic)
    # breakpoint()
    # print(rerun_entry_ids, delete_entry_ids)
    rerun_entry_ids = {
        ('dheu-4, dheu̯ə- (vermutlich: dhu̯ē-, vgl. die Erw. dhu̯ē-k-, dhu̯ē̆-s-)', '4. dheu-, dheu̯ə-, presumably dhu̯ē-, compare suffixes dhu̯ē-k-, dhu̯ē̆-s-'),
        ('gel-1', '1. gel-'),
        ('gēu-, gəu-, gū-', 'gēu-, gəu-, gū-'),
        ('e-3, ei-, i-, fem. ī-', '3. e-, ei-, i-, fem. ī-'),
        ('ger-3', '3. ger-'),
        ('er-3 : or- : r-', '3. er- : or- : r-')
    }
    # print out what needs to be rerun?
    redo_pokorny(rerun_entry_ids, filepath)
    # breakpoint()
    pass


if __name__ == '__main__':
    main()
